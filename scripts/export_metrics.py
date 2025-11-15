import os
import json
import csv
import re
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Color, PatternFill
from openpyxl.formatting.rule import ColorScaleRule

# Script to aggregate centralized and federated cross-day metrics into flat CSV/JSON exports.
# Output directory: exports/

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
CROSS_DAY_FED_DIR = os.path.join(ROOT, 'outputs', 'cross_day_fed')
CROSS_DAY_CENTRAL_DIR = os.path.join(ROOT, 'outputs', 'cross_day')
EXPORT_DIR = os.path.join(ROOT, 'exports')

os.makedirs(EXPORT_DIR, exist_ok=True)

Record = Dict[str, Any]

def parse_classification_report(report_dir: str) -> Dict[int, float]:
    """Parse sklearn classification_report.txt and return per-class f1 scores indexed by class int.
    Expects a file named classification_report.txt inside report_dir.
    Robust to varying whitespace. Lines starting with an integer class label are parsed.
    """
    path = os.path.join(report_dir, 'classification_report.txt')
    f1: Dict[int, float] = {}
    if not os.path.isfile(path):
        return f1
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or not line[0].isdigit():
                continue
            # Regex: class_label then multiple floats then support int
            # Example: '3     1.0000    0.4792    0.6479       144'
            m = re.match(r'^(\d+)\s+([0-9]*\.?[0-9]+)\s+([0-9]*\.?[0-9]+)\s+([0-9]*\.?[0-9]+)', line)
            if not m:
                continue
            cls = int(m.group(1))
            f1_val = float(m.group(4))
            f1[cls] = f1_val
    return f1


def parse_safety_metrics(report_dir: str) -> Dict[str, Optional[float]]:
    """Parse safety_metrics.txt extracting accuracy and critical error counts.
    Returns dict with keys: empty_accuracy, near_accuracy, mid_accuracy, far_accuracy,
    critical_near_as_empty, critical_near_as_empty_rate, false_alarm_empty_as_near,
    false_alarm_empty_as_near_rate. Missing values become None.
    """
    path = os.path.join(report_dir, 'safety_metrics.txt')
    keys = [
        'empty_accuracy', 'near_accuracy', 'mid_accuracy', 'far_accuracy',
        'critical_near_as_empty', 'critical_near_as_empty_rate',
        'false_alarm_empty_as_near', 'false_alarm_empty_as_near_rate'
    ]
    result: Dict[str, Optional[float]] = {k: None for k in keys}
    if not os.path.isfile(path):
        return result
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if ':' not in line:
                continue
            name, val = line.split(':', 1)
            name = name.strip()
            val = val.strip()
            if name in result:
                try:
                    result[name] = float(val)
                except ValueError:
                    result[name] = None
    return result


def collect_federated() -> List[Record]:
    records: List[Record] = []
    if not os.path.isdir(CROSS_DAY_FED_DIR):
        return records
    for entry in os.listdir(CROSS_DAY_FED_DIR):
        path = os.path.join(CROSS_DAY_FED_DIR, entry)
        if not os.path.isdir(path):
            continue
        # Expect folder pattern: model_dayX e.g. baseline_day1
        parts = entry.split('_day')
        if len(parts) != 2:
            continue
        model = parts[0]
        train_day_str = parts[1]
        if not train_day_str.isdigit():
            continue
        train_day = int(train_day_str)
        # inside: train_day{d}_{model}
        train_folder = f'train_day{train_day}_{model}'
        inner = os.path.join(path, train_folder)
        if not os.path.isdir(inner):
            continue
        cross_day_results = os.path.join(inner, 'cross_day_results.json')
        if os.path.isfile(cross_day_results):
            with open(cross_day_results, 'r', encoding='utf-8') as f:
                data = json.load(f)
            for k, v in data.get('results', {}).items():
                report_dir = os.path.join(inner, f'test_day{v["test_day"]}_report')
                f1_scores = parse_classification_report(report_dir)
                safety = parse_safety_metrics(report_dir)
                records.append({
                    'mode': 'federated',
                    'model': model,
                    'train_day': train_day,
                    'test_day': v['test_day'],
                    'accuracy': v['test_accuracy'],
                    'loss': v['test_loss'],
                    'num_samples': v.get('num_samples', None),
                    'rounds': data.get('rounds'),
                    'local_epochs': data.get('local_epochs'),
                    'normalize': data.get('normalize'),
                    **safety,
                    **{f'f1_class_{i}': f1_scores.get(i) for i in range(10)}
                })
        else:
            # fallback: individual test_day*_metrics.json files
            for fname in os.listdir(inner):
                if fname.startswith('test_day') and fname.endswith('_metrics.json'):
                    metrics_path = os.path.join(inner, fname)
                    with open(metrics_path, 'r', encoding='utf-8') as f:
                        v = json.load(f)
                    report_dir = os.path.join(inner, f'test_day{v["test_day"]}_report')
                    f1_scores = parse_classification_report(report_dir)
                    safety = parse_safety_metrics(report_dir)
                    records.append({
                        'mode': 'federated',
                        'model': model,
                        'train_day': train_day,
                        'test_day': v['test_day'],
                        'accuracy': v['test_accuracy'],
                        'loss': v['test_loss'],
                        'num_samples': v.get('num_samples', None),
                        'rounds': None,
                        'local_epochs': None,
                        'normalize': None,
                        **safety,
                        **{f'f1_class_{i}': f1_scores.get(i) for i in range(10)}
                    })
    return records

def collect_centralized() -> List[Record]:
    records: List[Record] = []
    if not os.path.isdir(CROSS_DAY_CENTRAL_DIR):
        return records
    # Folders may be like baseline_day0 or improved_day0? Inspect similar to federated? Use pattern search.
    for entry in os.listdir(CROSS_DAY_CENTRAL_DIR):
        path = os.path.join(CROSS_DAY_CENTRAL_DIR, entry)
        if not os.path.isdir(path):
            continue
        # Expect like baseline_day0 or improved_day0
        parts = entry.split('_day')
        if len(parts) != 2:
            continue
        model = parts[0]
        train_day_str = parts[1]
        if not train_day_str.isdigit():
            continue
        train_day = int(train_day_str)
        # inside: train_day{d}_{model}
        inner = os.path.join(path, f'train_day{train_day}_{model}')
        if not os.path.isdir(inner):
            continue
        for fname in os.listdir(inner):
            if fname.startswith('test_day') and fname.endswith('_metrics.json'):
                metrics_path = os.path.join(inner, fname)
                with open(metrics_path, 'r', encoding='utf-8') as f:
                    v = json.load(f)
                report_dir = os.path.join(inner, f'test_day{v["test_day"]}_report')
                f1_scores = parse_classification_report(report_dir)
                safety = parse_safety_metrics(report_dir)
                records.append({
                    'mode': 'centralized',
                    'model': model,
                    'train_day': train_day,
                    'test_day': v['test_day'],
                    'accuracy': v['test_accuracy'],
                    'loss': v['test_loss'],
                    'num_samples': v.get('num_samples', None),
                    'rounds': None,
                    'local_epochs': None,
                    'normalize': None,
                    **safety,
                    **{f'f1_class_{i}': f1_scores.get(i) for i in range(10)}
                })
    return records

def write_flat_csv(records: List[Record], out_path: str):
    if not records:
        return
    # Ensure consistent field ordering: basic keys then f1 scores
    safety_identifiers = [
        'empty_accuracy','near_accuracy','mid_accuracy','far_accuracy',
        'critical_near_as_empty','false_alarm_empty_as_near',
        'critical_near_as_empty_rate','false_alarm_empty_as_near_rate'
    ]
    base_keys = [k for k in records[0].keys() if k not in safety_identifiers and not k.startswith('f1_class_')]
    # Safety accuracy and counts (excluding rates) right after base keys
    safety_keys = [k for k in records[0].keys() if k in [
        'empty_accuracy','near_accuracy','mid_accuracy','far_accuracy','critical_near_as_empty','false_alarm_empty_as_near']]
    rate_keys = [k for k in records[0].keys() if k.endswith('_rate')]
    f1_keys = sorted([k for k in records[0].keys() if k.startswith('f1_class_')], key=lambda x: int(x.split('_')[-1]))
    fieldnames = base_keys + safety_keys + rate_keys + f1_keys
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in records:
            writer.writerow(r)


def write_json(records: List[Record], out_path: str):
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(records, f, indent=2)


def pivot_accuracy(records: List[Record], mode: str, model: str) -> Dict[str, Dict[str, float]]:
    # Returns nested dict: train_day -> test_day -> accuracy
    pivot: Dict[str, Dict[str, float]] = {}
    for r in records:
        if r['mode'] != mode or r['model'] != model:
            continue
        td = str(r['train_day'])
        if td not in pivot:
            pivot[td] = {}
        pivot[td][str(r['test_day'])] = r['accuracy'] * 100.0  # percent
    return pivot


def write_pivot_csv(pivot: Dict[str, Dict[str, float]], out_path: str):
    # Collect all test_days
    test_days = sorted({int(k2) for v in pivot.values() for k2 in v.keys()})
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        header = ['TrainDay'] + [f'TestDay{d}' for d in test_days]
        writer.writerow(header)
        for train_day in sorted(pivot.keys(), key=lambda x: int(x)):
            row = [train_day]
            for d in test_days:
                row.append(f'{pivot[train_day].get(str(d), ""):.2f}')
            writer.writerow(row)


def build_safety_summary(records: List[Record]) -> pd.DataFrame:
    rows = []
    for r in records:
        rows.append({
            'mode': r['mode'],
            'model': r['model'],
            'train_day': r['train_day'],
            'test_day': r['test_day'],
            'empty_accuracy': r.get('empty_accuracy'),
            'near_accuracy': r.get('near_accuracy'),
            'mid_accuracy': r.get('mid_accuracy'),
            'far_accuracy': r.get('far_accuracy'),
            'critical_near_as_empty': r.get('critical_near_as_empty'),
            'false_alarm_empty_as_near': r.get('false_alarm_empty_as_near')
        })
    df = pd.DataFrame(rows)
    # Aggregate: mean accuracies, sum critical errors per (mode, model, train_day)
    agg = df.groupby(['mode','model','train_day']).agg({
        'empty_accuracy':'mean',
        'near_accuracy':'mean',
        'mid_accuracy':'mean',
        'far_accuracy':'mean',
        'critical_near_as_empty':'sum',
        'false_alarm_empty_as_near':'sum'
    }).reset_index()
    agg = agg.sort_values(['mode','model','train_day'])
    return agg


def apply_conditional_formatting(ws, min_color="FFF5F0", mid_color="FDBE85", max_color="D94801"):
    # Applies a 3-color scale to numeric cells excluding header row
    max_col = ws.max_column
    max_row = ws.max_row
    if max_row < 2 or max_col < 2:
        return
    cell_range = f"B2:{ws.cell(row=max_row, column=max_col).coordinate}"
    rule = ColorScaleRule(start_type='min', start_color=min_color,
                          mid_type='percent', mid_value=50, mid_color=mid_color,
                          end_type='max', end_color=max_color)
    ws.conditional_formatting.add(cell_range, rule)


def write_excel(all_records: List[Record]):
    excel_path = os.path.join(EXPORT_DIR, 'metrics.xlsx')
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Flat sheet
        flat_df = pd.read_csv(os.path.join(EXPORT_DIR, 'metrics_flat.csv'))
        flat_df.to_excel(writer, sheet_name='flat', index=False)
        # Pivot sheets
        for mode in ['federated','centralized']:
            for model in ['baseline','improved']:
                pivot = pivot_accuracy(all_records, mode, model)
                if pivot:
                    # Convert pivot dict to DataFrame
                    df_rows = []
                    for train_day, cols in pivot.items():
                        row = {'TrainDay': int(train_day)}
                        for test_day, acc in cols.items():
                            row[f'TestDay{test_day}'] = acc
                        df_rows.append(row)
                    df = pd.DataFrame(df_rows).sort_values('TrainDay')
                    df.to_excel(writer, sheet_name=f'{mode}_{model}', index=False)
        # Safety summary
        safety_df = build_safety_summary(all_records)
        safety_df.to_excel(writer, sheet_name='safety_summary', index=False)
    # Reopen workbook for formatting
    wb = load_workbook(excel_path)
    # Apply formatting to pivot and safety sheets
    for sheet_name in wb.sheetnames:
        if sheet_name in ['flat']:
            continue
        ws = wb[sheet_name]
        apply_conditional_formatting(ws)
    wb.save(excel_path)
    return excel_path


def main():
    fed_records = collect_federated()
    cent_records = collect_centralized()
    all_records = fed_records + cent_records

    if not all_records:
        print('No records found. Ensure experiments were run.')
        return

    flat_csv = os.path.join(EXPORT_DIR, 'metrics_flat.csv')
    flat_json = os.path.join(EXPORT_DIR, 'metrics_flat.json')
    write_flat_csv(all_records, flat_csv)
    write_json(all_records, flat_json)

    for mode in ['federated', 'centralized']:
        for model in ['baseline', 'improved']:
            pivot = pivot_accuracy(all_records, mode, model)
            if pivot:
                pivot_csv = os.path.join(EXPORT_DIR, f'{mode}_{model}_accuracy_matrix.csv')
                write_pivot_csv(pivot, pivot_csv)

    # Write Excel workbook
    excel_path = write_excel(all_records)

    print('Export complete:')
    print(f'  Flat CSV: {flat_csv}')
    print(f'  Flat JSON: {flat_json}')
    print('  Matrices:')
    for mode in ['federated', 'centralized']:
        for model in ['baseline', 'improved']:
            path = os.path.join(EXPORT_DIR, f'{mode}_{model}_accuracy_matrix.csv')
            if os.path.isfile(path):
                print(f'    {path}')
    print(f'  Excel workbook: {excel_path}')

if __name__ == '__main__':
    main()
